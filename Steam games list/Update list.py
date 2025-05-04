import openpyxl
import requests
import time

def get_game_info(app_id):
    url = f"http://store.steampowered.com/api/appdetails/?appids={app_id}"
    response = requests.get(url)
    data = response.json()
    
    if data[str(app_id)]['success']:
        game_data = data[str(app_id)]['data']
        name = game_data['name']
        categories = [cat['description'] for cat in game_data.get('categories', [])]
        is_multiplayer = 'Multi-player' in categories
        return name, is_multiplayer
    
    return None, None

# Load the workbook and select the active sheet
workbook = openpyxl.load_workbook('Steam games list/Games List.xlsm')
sheet = workbook.active

# Assuming AppIDs are in column A, starting from row 2
for row in range(2, sheet.max_row + 1):
    app_id = sheet.cell(row=row, column=1).value
    if app_id:
        name, is_multiplayer = get_game_info(app_id)
        if name:
            sheet.cell(row=row, column=2, value=name)
            sheet.cell(row=row, column=3, value=str(is_multiplayer))
        print(f"Row: {row}, Name: {name}, AppID: {app_id}, IsMulti: {is_multiplayer}")
    
    time.sleep(1)  # Add a delay to avoid rate limiting

# Save the workbook
workbook.save('Steam games list/updated_excel_file.xlsx')